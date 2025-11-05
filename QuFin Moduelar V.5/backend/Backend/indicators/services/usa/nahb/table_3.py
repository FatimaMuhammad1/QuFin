from datetime import datetime
from typing import Iterable, Union, List
from pathlib import Path

import pyexcel
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.by import By

from indicators.countries import USA
from indicators.schemas import EntrySchema, IndicatorSchema
from indicators.services.common import make_wedriver, make_request_client
from indicators.utils import convert_to_float

_BASE_URL = (
    "https://www.nahb.org/news-and-economics/"
    "housing-economics/indices/housing-market-index"
)

_SINGLE_FAMILY_PRESENT = IndicatorSchema(
    country=USA,
    name="HMI - Single-Family: Present",
    code="HMI - Single-Family: Present",
    url="https://www.nahb.org/news-and-economics/housing-economics/indices/housing-market-index",
)
_SINGLE_FAMILY_NEX_SIX_MONTHS = IndicatorSchema(
    country=USA,
    name="HMI - Single-Family: Next Six Months",
    code="HMI - Single-Family: Next Six Months",
    url="https://www.nahb.org/news-and-economics/housing-economics/indices/housing-market-index",
)
_TRAFFIC_OF_PROSPECTIVE_BUYERS = IndicatorSchema(
    country=USA,
    name="HMI - Single-Family: Traffic of Prospective Buyers",
    code="HMI - Single-Family: Traffic of Prospective Buyers",
    url="https://www.nahb.org/news-and-economics/housing-economics/indices/housing-market-index",
)

INDICATORS: List[IndicatorSchema] = [
    _SINGLE_FAMILY_PRESENT,
    _SINGLE_FAMILY_NEX_SIX_MONTHS,
    _TRAFFIC_OF_PROSPECTIVE_BUYERS,
]


def _find_download_url(driver: WebDriver) -> Union[str, None]:
    anchor_tags = driver.find_elements(By.TAG_NAME, "a")

    url: Union[str, None] = None

    for anchor_tag in anchor_tags:
        label = anchor_tag.get_attribute("innerText")
        if "NAHB/Wells Fargo National HMI Components" in label:
            url = anchor_tag.get_attribute("href")
            break

    return url


def _download_excel_file(url: str) -> str:
    client = make_request_client()
    response = client.get(url, follow_redirects=True)

    path = Path("media/data")
    path.mkdir(exist_ok=True)
    filename = url.split("/")[-1]
    filepath = path / filename
    filepath.touch(exist_ok=True)

    with open(filepath, "wb") as file:
        file.write(response.content)

    return filepath


def _convert_to_xlsx(filepath: str) -> str:
    xlsx_filepath = f"{filepath}x"
    print(filepath)
    print(xlsx_filepath)

    pyexcel.save_book_as(file_name=str(filepath), dest_file_name=str(xlsx_filepath))

    return xlsx_filepath


def _load_worksheet(xlsx_filepath) -> Worksheet:
    wb = load_workbook(filename=xlsx_filepath)
    return wb.active  # type: ignore


def _extract_active_cols(worksheet: Worksheet) -> Iterable:
    current_year = datetime.now().year
    max_length = current_year - 1985 + 1
    return worksheet.iter_cols(min_col=2, max_col=max_length)  # type: ignore


def _extract_entries(
    cells: list,
    entries: List[EntrySchema],
    year_cell_index: int,
    row_start: int,
    row_end: int,
    indicator: IndicatorSchema,
) -> List[EntrySchema]:
    year = cells[year_cell_index].value
    prev_value = None

    for index, col_cell in enumerate(cells[row_start:row_end]):
        current_value = convert_to_float(col_cell.value)

        perc_change = 0.0

        if current_value is not None and prev_value not in [None, 0.0]:
            perc_change = round(
                (current_value - prev_value) / prev_value, ndigits=6  # type: ignore
            )  # type: ignore

        if current_value is not None:
            entry = EntrySchema(
                country=indicator.country,
                name=indicator.name,
                code=indicator.code,
                unit=indicator.unit,
                date=datetime(year=year, month=index + 1, day=1).date(),
                value1=current_value,
                perc_change1=perc_change,
            )
            entries.append(entry)

        prev_value = convert_to_float(col_cell.value)

    return entries


def get_entries() -> List[EntrySchema]:
    print("Making driver")
    driver: WebDriver = make_wedriver()
    print("Opening url")
    driver.get(_BASE_URL)
    print("Finding urls")
    url: Union[str, None] = _find_download_url(driver=driver)
    print("Quiting driver")
    driver.quit()

    if url is None:
        raise Exception("NAHB url cannot be found")

    filepath: str = _download_excel_file(url=url)
    xlsx_filepath: str = _convert_to_xlsx(filepath=filepath)

    worksheet: Worksheet = _load_worksheet(xlsx_filepath=xlsx_filepath)

    active_cols = _extract_active_cols(worksheet=worksheet)

    entries: List[EntrySchema] = []

    for col_cells in active_cols:
        # HMI - Single-Family: Present
        entries = _extract_entries(
            cells=col_cells,
            entries=entries,
            year_cell_index=3,
            row_start=4,
            row_end=16,
            indicator=_SINGLE_FAMILY_PRESENT,
        )

        # HMI - Single-Family: Next Six Months
        entries = _extract_entries(
            cells=col_cells,
            entries=entries,
            year_cell_index=18,
            row_start=19,
            row_end=31,
            indicator=_SINGLE_FAMILY_NEX_SIX_MONTHS,
        )

        # HMI - Single-Family: Traffic of Prospective Buyers
        entries = _extract_entries(
            cells=col_cells,
            entries=entries,
            year_cell_index=33,
            row_start=34,
            row_end=46,
            indicator=_TRAFFIC_OF_PROSPECTIVE_BUYERS,
        )

    return entries
